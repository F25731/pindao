from pathlib import Path
from typing import Any, Dict, List, Tuple, Generator, Sequence
from openpyxl import Workbook, load_workbook
import csv

CSV_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030")
CSV_HEADER_ALIASES = {
    "name": {"name", "名称", "资源名称", "标题", "title"},
    "tags": {"tags", "tag", "标签", "分类", "category"},
    "link": {"link", "链接", "资源链接", "源链接", "url", "share_url", "地址"},
}


def read_excel_stream(file_path: str, batch_size: int = 2000) -> Generator[List[Tuple[str, str, str]], None, None]:
    """
    流式读取 Excel 文件，每次 yield 一批行。
    适合大文件，不会一次性加载全部到内存。
    """
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    batch = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or len(row) < 3:
            continue
        name = str(row[0] or "").strip()
        tags = str(row[1] or "").strip()
        link = str(row[2] or "").strip()
        if name and link:
            batch.append((name, tags, link))
            if len(batch) >= batch_size:
                yield batch
                batch = []

    if batch:
        yield batch
    wb.close()


def read_csv_stream(file_path: str, batch_size: int = 2000) -> Generator[List[Tuple[str, str, str]], None, None]:
    """
    流式读取 CSV 文件，每次 yield 一批行。
    CSV 比 xlsx 快 10 倍以上，推荐百万级数据使用。
    """
    batch = []
    with _open_csv_file(file_path) as f:
        reader = _make_csv_reader(f)
        header = next(reader, None)
        indexes = _detect_column_indexes(header or [])
        for row in reader:
            if not row or len(row) < 3:
                continue
            name, tags, link = _extract_resource_columns(row, indexes)
            if name and link:
                batch.append((name, tags, link))
                if len(batch) >= batch_size:
                    yield batch
                    batch = []
    if batch:
        yield batch


def read_file_stream(file_path: str, batch_size: int = 2000) -> Generator[List[Tuple[str, str, str]], None, None]:
    """根据文件扩展名自动选择读取方式。"""
    ext = Path(file_path).suffix.lower()
    if ext == ".csv":
        yield from read_csv_stream(file_path, batch_size)
    else:
        yield from read_excel_stream(file_path, batch_size)


def _raw_payload(row: Sequence[Any], row_number: int) -> Dict[str, Any]:
    values = ["" if value is None else str(value).strip() for value in row]
    return {
        "row_number": row_number,
        "columns": values,
        "name": values[0] if len(values) > 0 else "",
        "tags": values[1] if len(values) > 1 else "",
        "link": values[2] if len(values) > 2 else "",
    }


def _open_csv_file(file_path: str):
    last_error = None
    for encoding in CSV_ENCODINGS:
        try:
            f = open(file_path, "r", encoding=encoding, newline="")
            f.read(4096)
            f.seek(0)
            return f
        except UnicodeDecodeError as exc:
            last_error = exc
            try:
                f.close()
            except Exception:
                pass
    raise last_error or UnicodeDecodeError("utf-8", b"", 0, 1, "无法识别 CSV 编码")


def _make_csv_reader(file_obj):
    sample = file_obj.read(8192)
    file_obj.seek(0)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel
    return csv.reader(file_obj, dialect)


def _clean_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("\ufeff", "")


def _detect_column_indexes(header: Sequence[Any]) -> Dict[str, int]:
    cleaned = [_clean_header(value) for value in header]
    indexes: Dict[str, int] = {}
    for key, aliases in CSV_HEADER_ALIASES.items():
        for index, value in enumerate(cleaned):
            if value in aliases:
                indexes[key] = index
                break

    indexes.setdefault("name", 0)
    indexes.setdefault("tags", 1)
    indexes.setdefault("link", 2)
    return indexes


def _extract_resource_columns(row: Sequence[Any], indexes: Dict[str, int]) -> Tuple[str, str, str]:
    values = ["" if value is None else str(value).strip() for value in row]

    def get_value(key: str) -> str:
        index = indexes.get(key, -1)
        if 0 <= index < len(values):
            return values[index]
        return ""

    return get_value("name"), get_value("tags"), get_value("link")


def read_csv_raw_stream(file_path: str, batch_size: int = 2000) -> Generator[List[Dict[str, Any]], None, None]:
    batch = []
    with _open_csv_file(file_path) as f:
        reader = _make_csv_reader(f)
        header = next(reader, None)
        indexes = _detect_column_indexes(header or [])
        for row_number, row in enumerate(reader, start=2):
            payload = _raw_payload(row, row_number)
            name, tags, link = _extract_resource_columns(row, indexes)
            payload["name"] = name
            payload["tags"] = tags
            payload["link"] = link
            if len(batch) >= batch_size:
                yield batch
                batch = []
            batch.append(payload)
    if batch:
        yield batch


def read_excel_raw_stream(file_path: str, batch_size: int = 2000) -> Generator[List[Dict[str, Any]], None, None]:
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    batch = []
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_number == 1:
            continue
        batch.append(_raw_payload(row or [], row_number))
        if len(batch) >= batch_size:
            yield batch
            batch = []
    if batch:
        yield batch
    wb.close()


def read_file_raw_stream(file_path: str, batch_size: int = 2000) -> Generator[List[Dict[str, Any]], None, None]:
    ext = Path(file_path).suffix.lower()
    if ext == ".csv":
        yield from read_csv_raw_stream(file_path, batch_size)
    else:
        yield from read_excel_raw_stream(file_path, batch_size)


def write_excel(
    file_path: str,
    rows: List[Sequence[str]],
    headers: Sequence[str] = ("名称", "标签", "链接"),
):
    """写入 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "资源列表"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    widths = [40, 20, 65, 65, 24, 24]
    for index, width in enumerate(widths[:len(headers)], start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    wb.save(file_path)
